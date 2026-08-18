"""
Commodity Option Valuator Pro
=============================

Market Data File Reader.

Commit 0014
------------

Read option market data exported from external
market-data applications and normalize it into
records consumable by MarketDataAdapter.

Author : Simon
Version : 0.4.0
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


class MarketFileReader:
    """
    Read option market data files.

    Supported input
    ---------------
    - Tab-separated text files
    - CSV-like files
    - XLS files exported as text
    - GBK / UTF-8 encoded files

    The reader intentionally returns dictionaries
    rather than OptionContract objects.

    Conversion into domain objects remains the
    responsibility of MarketDataAdapter.
    """

    DEFAULT_ENCODINGS = (
        "utf-8-sig",
        "utf-8",
        "gbk",
        "gb18030",
    )

    REQUIRED_FIELDS = (
        "symbol",
        "direction",
        "strike",
        "price",
        "volume",
    )

    FIELD_ALIASES: dict[str, tuple[str, ...]] = {
        "symbol": (
            "symbol",
            "合约代码",
            "期权代码",
            "证券代码",
            "代码",
        ),
        "direction": (
            "direction",
            "方向",
            "期权方向",
            "类型",
        ),
        "strike": (
            "strike",
            "行权价",
            "执行价",
            "行权价格",
        ),
        "price": (
            "price",
            "价格",
            "市场价格",
            "最新价",
            "最新价格",
            "结算价",
        ),
        "volume": (
            "volume",
            "成交量",
            "成交",
        ),
        "open_interest": (
            "open_interest",
            "持仓量",
            "持仓",
            "未平仓量",
        ),
        "bid": (
            "bid",
            "买价",
            "买一价",
            "买入价",
        ),
        "ask": (
            "ask",
            "卖价",
            "卖一价",
            "卖出价",
        ),
    }

    def __init__(
        self,
        encodings: tuple[str, ...] | None = None,
    ) -> None:
        """
        Initialize the file reader.

        Parameters
        ----------
        encodings:
            Encoding candidates used when reading text files.
        """

        self.encodings = (
            encodings
            if encodings is not None
            else self.DEFAULT_ENCODINGS
        )

    # ======================================================
    # Public API
    # ======================================================

    def read(
        self,
        path: str | Path,
    ) -> list[dict[str, Any]]:
        """
        Read a market-data file.

        Parameters
        ----------
        path:
            File path.

        Returns
        -------
        list[dict[str, Any]]
            Raw normalized records.

        Raises
        ------
        FileNotFoundError
            If the file does not exist.

        ValueError
            If the file is empty or its header cannot
            be interpreted.
        """

        file_path = Path(path)

        if not file_path.exists():
            raise FileNotFoundError(
                f"市场数据文件不存在：{file_path}"
            )

        if not file_path.is_file():
            raise ValueError(
                f"市场数据路径不是文件：{file_path}"
            )

        suffix = file_path.suffix.lower()

        if suffix in {
            ".xlsx",
            ".xlsm",
        }:
            return self._read_excel(
                file_path
            )

        return self._read_text(
            file_path
        )

    # ======================================================
    # Text Reader
    # ======================================================

    def _read_text(
        self,
        path: Path,
    ) -> list[dict[str, Any]]:
        """Read a delimited text file."""

        text = self._read_text_content(
            path
        )

        lines = [
            line.strip()
            for line in text.splitlines()
            if line.strip()
        ]

        if not lines:
            raise ValueError(
                "市场数据文件为空。"
            )

        delimiter = self._detect_delimiter(
            lines[0]
        )

        headers = [
            item.strip()
            for item in lines[0].split(
                delimiter
            )
        ]

        normalized_headers = (
            self._normalize_headers(
                headers
            )
        )

        self._validate_headers(
            normalized_headers
        )

        records: list[dict[str, Any]] = []

        for line_number, line in enumerate(
            lines[1:],
            start=2,
        ):
            values = [
                item.strip()
                for item in line.split(
                    delimiter
                )
            ]

            if not any(values):
                continue

            if len(values) < len(headers):
                values.extend(
                    [""] * (
                        len(headers)
                        - len(values)
                    )
                )

            record = {
                normalized_headers[index]: (
                    values[index]
                )
                for index in range(
                    len(headers)
                )
            }

            records.append(
                self._normalize_record(
                    record
                )
            )

        return records

    def _read_text_content(
        self,
        path: Path,
    ) -> str:
        """Read text using configured encoding candidates."""

        last_error: UnicodeDecodeError | None = None

        for encoding in self.encodings:

            try:
                return path.read_text(
                    encoding=encoding
                )

            except UnicodeDecodeError as exc:
                last_error = exc

        if last_error is not None:
            raise ValueError(
                f"无法识别市场数据文件编码：{path}"
            ) from last_error

        raise ValueError(
            f"无法读取市场数据文件：{path}"
        )

    # ======================================================
    # Excel Reader
    # ======================================================

    def _read_excel(
        self,
        path: Path,
    ) -> list[dict[str, Any]]:
        """
        Read an XLSX/XLSM workbook.

        The reader uses openpyxl for modern Excel files.
        """

        try:
            from openpyxl import load_workbook

        except ImportError as exc:
            raise RuntimeError(
                "读取 Excel 文件需要安装 openpyxl。"
            ) from exc

        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )

        try:
            worksheet = workbook.active

            rows = list(
                worksheet.iter_rows(
                    values_only=True
                )
            )

        finally:
            workbook.close()

        if not rows:
            raise ValueError(
                "市场数据 Excel 文件为空。"
            )

        headers = [
            self._stringify(value)
            for value in rows[0]
        ]

        normalized_headers = (
            self._normalize_headers(
                headers
            )
        )

        self._validate_headers(
            normalized_headers
        )

        records: list[dict[str, Any]] = []

        for row in rows[1:]:

            if not any(
                value is not None
                and str(value).strip()
                for value in row
            ):
                continue

            values = list(row)

            if len(values) < len(headers):
                values.extend(
                    [None] * (
                        len(headers)
                        - len(values)
                    )
                )

            record = {
                normalized_headers[index]: (
                    values[index]
                )
                for index in range(
                    len(headers)
                )
            }

            records.append(
                self._normalize_record(
                    record
                )
            )

        return records

    # ======================================================
    # Header Processing
    # ======================================================

    def _normalize_headers(
        self,
        headers: list[str],
    ) -> list[str]:
        """Map external column names to internal names."""

        normalized: list[str] = []

        for header in headers:

            clean = self._clean_header(
                header
            )

            mapped = self._map_header(
                clean
            )

            normalized.append(
                mapped
            )

        return normalized

    def _map_header(
        self,
        header: str,
    ) -> str:
        """Map one external header to a canonical field."""

        for field, aliases in (
            self.FIELD_ALIASES.items()
        ):

            for alias in aliases:

                if header == self._clean_header(
                    alias
                ):
                    return field

        return header

    @staticmethod
    def _clean_header(
        value: str,
    ) -> str:
        """Normalize a header for matching."""

        return (
            str(value)
            .strip()
            .replace(
                " ",
                "",
            )
            .replace(
                "\t",
                "",
            )
            .lower()
        )

    def _validate_headers(
        self,
        headers: list[str],
    ) -> None:
        """Validate required market-data columns."""

        missing = [
            field
            for field in self.REQUIRED_FIELDS
            if field not in headers
        ]

        if missing:
            raise ValueError(
                "市场数据文件缺少必要字段："
                + "、".join(missing)
            )

    # ======================================================
    # Record Processing
    # ======================================================

    def _normalize_record(
        self,
        record: dict[str, Any],
    ) -> dict[str, Any]:
        """
        Normalize one market-data record.

        Values remain mostly untouched so that
        MarketDataAdapter remains responsible for
        domain-level validation and conversion.
        """

        normalized = dict(record)

        for field in (
            "symbol",
            "direction",
        ):
            if field in normalized:
                normalized[field] = (
                    self._stringify(
                        normalized[field]
                    ).strip()
                )

        return normalized

    # ======================================================
    # Helpers
    # ======================================================

    @staticmethod
    def _detect_delimiter(
        header: str,
    ) -> str:
        """Detect common text-file delimiters."""

        candidates = (
            "\t",
            ",",
            ";",
            "|",
        )

        best = "\t"
        best_count = -1

        for delimiter in candidates:

            count = header.count(
                delimiter
            )

            if count > best_count:
                best = delimiter
                best_count = count

        return best

    @staticmethod
    def _stringify(
        value: Any,
    ) -> str:
        """Convert a cell value to string safely."""

        if value is None:
            return ""

        return str(value)

    # ======================================================
    # Public Utility API
    # ======================================================

    def read_records(
        self,
        path: str | Path,
    ) -> list[dict[str, Any]]:
        """
        Alias for read().

        Kept for readability at integration boundaries.
        """

        return self.read(
            path
        )


__all__ = [
    "MarketFileReader",
]